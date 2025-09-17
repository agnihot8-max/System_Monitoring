import psutil
import time
import pandas as pd
from datetime import datetime
from collections import defaultdict
from openpyxl.chart import LineChart, Reference


LOG_INTERVAL_SECONDS = 5
RUN_DURATION_SECONDS = 60.0
CPU_ALERT_THRESHOLD = 80.0
OUTPUT_FILENAME = 'system_monitor_report.xlsx'
THREAT_LOG_FILENAME = 'cpu_threat_log.txt'



def get_top_processes(num_processes=3):
    processes = []
    for proc in psutil.process_iter(['pid', 'name', 'cpu_percent']):
        try:
            proc.cpu_percent(interval=0.01)
            processes.append(proc)
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass
    time.sleep(0.1)
    top_processes = []
    for proc in processes:
        try:
            cpu_usage = proc.cpu_percent()
            if cpu_usage > 0.0:
                top_processes.append({'pid': proc.info['pid'], 'name': proc.info['name'], 'cpu': cpu_usage})
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass
    sorted_processes = sorted(top_processes, key=lambda p: p['cpu'], reverse=True)
    output = "\n  Top Processes:"
    count = 0
    for p in sorted_processes:
        if p['name'] not in ['System Idle Process', 'System']:
            output += f" {p['name']}({p['cpu']:.1f}%)"
            count += 1
            if count >= num_processes:
                break
    return output


def get_network_connections():
    connections = psutil.net_connections(kind='inet')
    established_conns = defaultdict(int)
    total_established = 0
    for conn in connections:
        if conn.status == 'ESTABLISHED' and conn.raddr:
            established_conns[conn.raddr[0]] += 1
            total_established += 1
    output = f"\n  Network: {total_established} established connections."
    if established_conns:
        sorted_conns = sorted(established_conns.items(), key=lambda item: item[1], reverse=True)
        conn_details = ", ".join([f"{ip}({count})" for ip, count in sorted_conns[:3]])
        output += f" Connected To: {conn_details}"
    return output


def add_chart_to_excel(writer):
    workbook = writer.book
    worksheet = writer.sheets['System_Log']
    chart_sheet = workbook.create_sheet('CPU_Graph')

    chart = LineChart()
    chart.title = "CPU Usage Over Time"
    chart.x_axis.title = "Time"
    chart.y_axis.title = "CPU Usage (%)"


    data = Reference(worksheet, min_col=2, min_row=1, max_col=2, max_row=worksheet.max_row)
    categories = Reference(worksheet, min_col=1, min_row=2, max_row=worksheet.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart_sheet.add_chart(chart, "A1")


def main():
    start_time = time.time()
    data_log = []
    alert_log = []

    print("Starting system monitoring...")
    try:
        while True:
            if RUN_DURATION_SECONDS and (time.time() - start_time) > RUN_DURATION_SECONDS:
                print("\nMonitoring duration reached. Exiting.")
                break

            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            cpu_usage = psutil.cpu_percent(interval=None)
            memory_info = psutil.virtual_memory()

            current_data = {
                'Timestamp': timestamp,
                'CPU Usage (%)': cpu_usage,
                'Memory Usage (%)': memory_info.percent
            }
            data_log.append(current_data)

            top_processes_str = get_top_processes()
            network_connections_str = get_network_connections()

            log_output = (
                f"\n[{timestamp}] | CPU: {cpu_usage}% | Memory: {memory_info.percent}%"
            )
            print(log_output + network_connections_str + top_processes_str)

            if cpu_usage > CPU_ALERT_THRESHOLD:
                alert_msg = f"High CPU usage detected: {cpu_usage}%"
                print(f"ALERT: {alert_msg}")

                alert_log.append({
                    'Timestamp': timestamp,
                    'AlertType': 'High CPU Usage',
                    'Details': alert_msg
                })

                with open(THREAT_LOG_FILENAME, 'a') as f:
                    f.write(f"[{timestamp}] - {alert_msg}\n")

            time.sleep(LOG_INTERVAL_SECONDS)

    except KeyboardInterrupt:
        print("\nMonitoring stopped by user.")
    finally:
        if data_log:
            print(f"\nSaving data to {OUTPUT_FILENAME}...")
            with pd.ExcelWriter(OUTPUT_FILENAME, engine='openpyxl') as writer:
                df_data = pd.DataFrame(data_log)
                df_data.to_excel(writer, sheet_name='System_Log', index=False)

                if alert_log:
                    df_alerts = pd.DataFrame(alert_log)
                    df_alerts.to_excel(writer, sheet_name='Alert_Tickets', index=False)

                add_chart_to_excel(writer)

            print("Report saved successfully.")
        else:
            print("No data was logged.")


if __name__ == "__main__":
    main()
